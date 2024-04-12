import pandas as pd
import openpyxl
from openpyxl import load_workbook,Workbook
from datetime import datetime,timedelta
from openpyxl.styles import *
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
def reduce_excel_col_name(n):
    # initialize output string as empty
    result = ''

    while n > 0:
        # find the index of the next letter and concatenate the letter
        # to the solution

        # here index 0 corresponds to 'A', and 25 corresponds to 'Z'
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26

    return result[::-1]
def formatting():
    global font_bold,font_notbold,border,purpleFill,yellowFill,orangeFill,no_fill,no_border,GrayFill,GreenFill
    no_fill = PatternFill(fill_type=None)
    side = Side(border_style=None)
    no_border = borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    font_bold=Font(name="Arial",size=11,bold=True,color="000000")
    font_notbold=Font(name="Arial",size=11,bold=False,color="000000")
    thin=Side(border_style="thin", color="000000")
    double=Side(border_style="double", color="000000")
    border=Border(top=thin, left=thin, right=thin, bottom=thin)

    yellow="FFFF99"  ##YELLOW
    yellowFill=PatternFill(start_color=yellow,
                             end_color=yellow,
                             fill_type='solid')

    purple='9999FF'
    purpleFill=PatternFill(start_color=purple,
                             end_color=purple,
                             fill_type='solid')

    orange="FFC125"  ##orange
    OrangeFill=PatternFill(start_color=orange,
                             end_color=orange,
                             fill_type='solid')
    gray="E7E6E6"  ##gray
    GrayFill=PatternFill(start_color=gray,
                             end_color=gray,
                             fill_type='solid')
    green="B8C2AD"  ##green

    GreenFill=PatternFill(start_color=green,
                             end_color=green,
                             fill_type='solid')

def worksheet(ws,header,semi_attribution_ytd,number):
    for i in range(len(header)):
        ws[reduce_excel_col_name(i+1)+str(4)].value = header[i]
        ws[reduce_excel_col_name(i + 1) + str(4)].alignment=Alignment(vertical="bottom",horizontal="center")
        ws[reduce_excel_col_name(i + 1) + str(4)].fill=GreenFill
        ws[reduce_excel_col_name(i + 1) + str(4)].font = font_bold

    rows = dataframe_to_rows(semi_attribution_ytd, index=False, header=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+4, column=c_idx, value=value)
            ws[reduce_excel_col_name(c_idx)+str(r_idx+4)].font=font_notbold
    for i in range(5,r_idx+4+1):
        for j in range(4):
            ws[reduce_excel_col_name(3+j)+str(i)].value = float(ws[reduce_excel_col_name(3+j)+str(i)].value)/100
            if j !=3:
                ws[reduce_excel_col_name(3 + j) + str(i)].number_format="0.0%"
            else:
                ws[reduce_excel_col_name(3 + j) + str(i)].number_format = "0.00%"
    ws["A"+str(r_idx+4+7)].value = "SVLO SEMI (SINCE 20230616)"
    ws["C"+str(r_idx+4+7)].value = "=SUM(C5:C"+str(r_idx+4)+")"
    ws["D" + str(r_idx + 4 + 7)].value= "=SUM(D5:D"+str(r_idx+4)+")"
    ws["F" + str(r_idx + 4 + 7)].value = "=SUM(F5:F" + str(r_idx + 4) + ")"

    ws["A"+str(r_idx+4+7)].font=font_notbold
    ws["C"+str(r_idx+4+7)].font=font_notbold
    ws["D" + str(r_idx + 4 + 7)].font=font_notbold
    ws["F" + str(r_idx + 4 + 7)].font = font_notbold

    ws["C"+str(r_idx+4+7)].number_format="0.0%"
    ws["D" + str(r_idx + 4 + 7)].font=number_format="0.0%"
    ws["F" + str(r_idx + 4 + 7)].font = number_format = "0.0%"
    if number ==1:
        header_region=["Region","% Avg Portfolio Weight","YTD "+str((datetime.today().replace(day=1)-timedelta(1)).year)+" Attribution"]
    elif number ==0:
        header_region = ["Region", "% Avg Portfolio Weight", "Attribution since 0616"]
    for i in range(len(header_region)):
        ws[reduce_excel_col_name(12+i)+str(4)].value = header_region[i]
        if i!=0:
            ws[reduce_excel_col_name(12 + i)+str(4)].alignment=Alignment(horizontal="center",vertical="center")
        else:
            ws[reduce_excel_col_name(12 + i)+str(4)].alignment=Alignment(horizontal="left",vertical="bottom")
        ws[reduce_excel_col_name(12 + i)+str(4)].font=font_bold
        ws[reduce_excel_col_name(12 + i) + str(4)].fill=GreenFill
    region_list=semi_attribution_ytd.groupby(['Region'])["Contribution to Return (%)"].sum().sort_values(ascending=False).index
    for i in range(len(region_list)):
        ws["L"+str(5+i)].value=region_list[i]
        ws["M"+str(5+i)].value = "=SUMIF(G5:G"+str(r_idx + 4)+",L"+str(5+i)+",C5:C"+str(r_idx + 4)+")"
        ws["N"+str(5+i)].value = "=SUMIF(G5:G"+str(r_idx + 4)+",L"+str(5+i)+",F5:F"+str(r_idx + 4)+")"

        ws["L"+str(5+i)].font=font_notbold
        ws["M"+str(5+i)].font=font_notbold
        ws["N"+str(5+i)].font=font_notbold

        ws["L"+str(5+i)].alignment = Alignment(horizontal="left",vertical="bottom")
        ws["M"+str(5+i)].alignment=Alignment(horizontal="center",vertical="center")
        ws["N"+str(5+i)].alignment=Alignment(horizontal="center",vertical="center")

        ws["M"+str(5+i)].number_format = "0.0%"
        ws["N"+str(5+i)].number_format = "0.0%"
    number_row=r_idx + 4
    if number==1:
        ws["L"+str(5+i+4)].value = "Top 10 Winners (YTD"+str(lastmonth_lastday.year)+")"
    elif number == 0 :
        ws["L" + str(5 + i + 4)].value = "Top 10 Winners"
    ws["L" + str(5 + i + 4)].font= Font(name="Arial",size=11,bold=True,color="000000" , underline="single")
    semi_attribution_ytd.columns =  header
    semi_attribution_ytd=semi_attribution_ytd.sort_values(by=["USD Attribution (%)"],ascending=False)
    top10_ytd = semi_attribution_ytd.loc[semi_attribution_ytd["USD Attribution (%)"]>0].iloc[:10]
    bottom10_ytd = semi_attribution_ytd.iloc[-10:]
    bottom10_ytd=bottom10_ytd[bottom10_ytd["USD Attribution (%)"]<0].sort_values(by=["USD Attribution (%)"],ascending=True)
    header_10 = ["Ticker", "Name", "% Avg Portfolio Weight", "USD Total Return (%)",
              "USD Attribution (%)", "Region", "GICS Sub-Industry Name", "Theme"]
    top10_ytd=top10_ytd[header_10]
    bottom10_ytd=bottom10_ytd[header_10]
    top10_row=5+i+4
    rows = dataframe_to_rows(top10_ytd, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+top10_row, column=c_idx+11, value=value)
            ws[reduce_excel_col_name(c_idx+11)+str(r_idx+top10_row)].font=font_notbold
            #ws[reduce_excel_col_name(c_idx)+str(r_idx+top10_row)].font=font_notbold

    for i in range(top10_row+2,r_idx+top10_row+1):
        ws["N"+str(i)].value = float(ws["N"+str(i)].value)/100
        ws["O" + str(i)].value = float(ws["O" + str(i)].value) / 100
        ws["P" + str(i)].value = float(ws["P" + str(i)].value) / 100
        ws["N" + str(i)].number_format="0.0%"
        ws["O" + str(i)].number_format="0.0%"
        ws["P" + str(i)].number_format="0.0%"


    bottom10_row = r_idx+top10_row + 2
    if number ==1 :
        ws["L"+str(bottom10_row)].value = "Bottom 10 (YTD "+str(lastmonth_lastday.year)+")"
    elif number ==0:
        ws["L" + str(bottom10_row)].value = "Bottom 10"
    ws["L" + str(bottom10_row)].font= Font(name="Arial",size=11,bold=True,color="000000" , underline="single")
    rows = dataframe_to_rows(bottom10_ytd, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+bottom10_row, column=c_idx+11, value=value)
            ws[reduce_excel_col_name(c_idx+11)+str(r_idx+bottom10_row)].font=font_notbold
    for i in range(bottom10_row+2,r_idx+bottom10_row+1):
        ws["N"+str(i)].value = ws["N"+str(i)].value/100
        ws["O" + str(i)].value = ws["O" + str(i)].value / 100
        ws["P" + str(i)].value = ws["P" + str(i)].value / 100
        ws["N" + str(i)].number_format="0.0%"
        ws["O" + str(i)].number_format="0.0%"
        ws["P" + str(i)].number_format="0.0%"

    for  i in range(len(header_10)):
        ws[reduce_excel_col_name(12+i)+str(top10_row+1)].font=font_bold
        ws[reduce_excel_col_name(12+i)+str(bottom10_row+1)].font=font_bold

        ws[reduce_excel_col_name(12+i)+str(top10_row+1)].fill=GreenFill
        ws[reduce_excel_col_name(12+i)+str(bottom10_row+1)].fill=GreenFill
        if i ==0:
            ws[reduce_excel_col_name(12 + i) + str(top10_row + 1)].alignment = Alignment(horizontal="left",vertical="bottom")
            ws[reduce_excel_col_name(12 + i) + str(bottom10_row + 1)].alignment = Alignment(horizontal="left",vertical="bottom")
        else:
            ws[reduce_excel_col_name(12 + i) + str(top10_row + 1)].alignment = Alignment(horizontal="center" ,vertical="bottom")
            ws[reduce_excel_col_name(12 + i) + str(bottom10_row + 1)].alignment = Alignment(horizontal="center",vertical="bottom")
    GICS_row=r_idx+bottom10_row+8
    GICS_header=["GICS Sub-Industry","% Avg Portfolio Weight","Attribution since 0616"]
    for i in range(len(GICS_header)):
        ws[reduce_excel_col_name(12+i)+str(GICS_row)].value = GICS_header[i]
        ws[reduce_excel_col_name(12 + i) + str(GICS_row)].font=font_bold
        ws[reduce_excel_col_name(12 + i) + str(GICS_row)].fill=GreenFill
        if i ==0:
            ws[reduce_excel_col_name(12 + i) + str(GICS_row)].alignment=Alignment(vertical="center",horizontal="left")
        else:
            ws[reduce_excel_col_name(12 + i) + str(GICS_row)].alignment = Alignment(vertical="center",horizontal="center")
    for i in range(len(semi_attribution_ytd["GICS Sub-Industry Name"].unique())):
        ws["L"+str(GICS_row+1+i)].value=semi_attribution_ytd["GICS Sub-Industry Name"].unique()[i]
        ws["L" + str(GICS_row + 1 + i)].alignment=Alignment(vertical="bottom",horizontal="left")
        ws["L" + str(GICS_row + 1 + i)].font=font_notbold

        ws["M"+str(GICS_row + 1 + i)].value = "=SUMIF(H5:H"+str(number_row)+",L"+str(GICS_row + 1 + i)+",C5:C"+str(r_idx + 4)+")"
        ws["N"+str(GICS_row + 1 + i)].value = "=SUMIF(H5:H"+str(number_row)+",L"+str(GICS_row + 1 + i)+",F5:F"+str(r_idx + 4)+")"
        ws["M" + str(GICS_row + 1 + i)].font=font_notbold
        ws["N" + str(GICS_row + 1 + i)].font=font_notbold
        ws["M" + str(GICS_row + 1 + i)].alignment=Alignment(vertical="bottom",horizontal="center")
        ws["N" + str(GICS_row + 1 + i)].alignment=Alignment(vertical="bottom",horizontal="center")
        ws["M" + str(GICS_row + 1 + i)].number_format="0.0%"
        ws["N" + str(GICS_row + 1 + i)].number_format="0.0%"
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    return ws


if __name__ == "__main__":
    formatting()
    lastmonth_lastday=datetime.today().replace(day=1)-timedelta(1)
    lastyear_lastday=datetime(lastmonth_lastday.year-1,12,31)
    semi_attribution_ytd_file="SVLO Global Semi Attribution Summary ("+lastyear_lastday.strftime("%y%m%d")+"-"+lastmonth_lastday.strftime("%y%m%d")+")_raw.xlsx"

    semi_attribution_ytd=pd.read_excel(semi_attribution_ytd_file,sheet_name="Attribution",header=11)
    semi_attribution_ytd=semi_attribution_ytd.iloc[1:-2,1:-4]
    semi_attribution_ytd = semi_attribution_ytd.drop('% Wgt Chg', axis=1)
    semi_attribution_ytd=semi_attribution_ytd.sort_values(by=["% Average Weight"],ascending=False)
    semi_attribution_ytd.rename(columns = {semi_attribution_ytd.columns[0]:"Name"},inplace = True)
    ticker_ytd = semi_attribution_ytd["Ticker"]
    name_ytd = semi_attribution_ytd["Name"]
    semi_attribution_ytd.rename(columns={semi_attribution_ytd.columns[0]: "Ticker",semi_attribution_ytd.columns[1]: "Name"}, inplace=True)
    semi_attribution_ytd["Ticker"]=ticker_ytd
    semi_attribution_ytd["Name"]=name_ytd


    paper_portfolio = pd.read_excel("SVLO Semi Paper Portfolio_" + lastmonth_lastday.strftime("%Y%m%d") + ".xlsx",sheet_name="Portfolio Weighting", header=3)
    paper_portfolio=paper_portfolio.dropna(subset=["Ticker"],axis=0)
    paper_portfolio["Ticker"]=[i.split(" Equity")[0] for i in paper_portfolio["Ticker"]]
    region=paper_portfolio[["Ticker","Country"]]
    region=region.dropna()
    region_ordered=[]
    #region_ordered = [region[region[region.columns[0]] == i][region.columns[1]].item() if i in list(region[region.columns[0]]) else "" for i in (ticker_ytd)]
    for i in (ticker_ytd):
        if i.split(" ")[1] == "US":
            region_ordered.append("United States")
        elif i.split(" ")[1] == "NA":
            region_ordered.append("Netherlands")
        elif i.split(" ")[1] == "JP":
            region_ordered.append("Japan")
        elif i.split(" ")[1] == "TT":
            region_ordered.append("Taiwan")
        elif i.split(" ")[1] == "KS":
            region_ordered.append("Korea")
        elif i.split(" ")[1] == "GR":
            region_ordered.append("Germany")
        elif i.split(" ")[1] == "CH":
            region_ordered.append("China")
        elif i.split(" ")[1] == "FR":
            region_ordered.append("France")
        else:
            region_ordered.append("Others")

    theme=paper_portfolio[["Ticker","Theme"]]
    theme=theme.dropna()
    theme[theme.columns[0]]=[i.split(" Equity")[0] for i in theme[theme.columns[0]]]
    GICS=paper_portfolio[["Ticker","GICS Sub-Industry Name"]]

    theme_ordered=[theme[theme[theme.columns[0]]==i][theme.columns[1]].item() if i in list(theme[theme.columns[0]]) else "" for i in (ticker_ytd)]
    GICS_ordered = [GICS[GICS[GICS.columns[0]] == i][GICS.columns[1]].item() if i in list(GICS[GICS.columns[0]]) else "" for i in (ticker_ytd)]
    semi_attribution_ytd["Region"]=region_ordered
    semi_attribution_ytd["GICS"]=GICS_ordered
    semi_attribution_ytd["Theme"]=theme_ordered

    wb=Workbook()
    wb.create_sheet("Attribution YTD "+str(lastmonth_lastday.year))
    wb.create_sheet('Attribution ITD')
    header= ["Ticker","Name","% Avg Portfolio Weight","Portfolio Weight as of "+lastmonth_lastday.strftime("%Y/%m/%d"),"USD Total Return (%)","USD Attribution (%)","Region","GICS Sub-Industry Name","Theme"]
    ws=wb["Attribution YTD "+str(lastmonth_lastday.year)]
    ws["A1"].value= "PORT Attribution Report: ("+(lastyear_lastday+timedelta(1)).strftime("%Y/%m/%d")+" - "+lastmonth_lastday.strftime("%Y/%m/%d")+")"
    ws["A1"].font=Font(name="Arial",size=12,bold=True,color="000000")
    ws=worksheet(ws,header,semi_attribution_ytd,1)

#sheet ITD

    semi_attribution_itd_file="SVLO Global Semi Attribution Summary (230616-"+lastmonth_lastday.strftime("%y%m%d")+")_raw.xlsx"

    semi_attribution_itd=pd.read_excel(semi_attribution_itd_file,sheet_name="Attribution",header=11)
    semi_attribution_itd=semi_attribution_itd.iloc[1:-2,1:-4]
    print(semi_attribution_itd.columns)
    semi_attribution_itd = semi_attribution_itd.drop('% Wgt Chg', axis=1)
    semi_attribution_itd=semi_attribution_itd.sort_values(by=["% Average Weight"],ascending=False)
    semi_attribution_itd.rename(columns = {semi_attribution_itd.columns[0]:"Name"},inplace = True)
    ticker_ytd = semi_attribution_itd["Ticker"]
    name_ytd = semi_attribution_itd["Name"]
    semi_attribution_itd.rename(columns={semi_attribution_itd.columns[0]: "Ticker",semi_attribution_itd.columns[1]: "Name"}, inplace=True)
    semi_attribution_itd["Ticker"]=ticker_ytd
    semi_attribution_itd["Name"]=name_ytd


    paper_portfolio = pd.read_excel("SVLO Semi Paper Portfolio_" + lastmonth_lastday.strftime("%Y%m%d") + ".xlsx",sheet_name="Portfolio Weighting", header=3)
    paper_portfolio=paper_portfolio.dropna(subset=["Ticker"],axis=0)
    paper_portfolio["Ticker"]=[i.split(" Equity")[0] for i in paper_portfolio["Ticker"]]
    region=paper_portfolio[["Ticker","Country"]]
    region=region.dropna()
    region_ordered=[]
    #region_ordered = [region[region[region.columns[0]] == i][region.columns[1]].item() if i in list(region[region.columns[0]]) else "" for i in (ticker_ytd)]
    for i in (ticker_ytd):
        if i.split(" ")[1] == "US":
            region_ordered.append("United States")
        elif i.split(" ")[1] == "NA":
            region_ordered.append("Netherlands")
        elif i.split(" ")[1] == "JP":
            region_ordered.append("Japan")
        elif i.split(" ")[1] == "TT":
            region_ordered.append("Taiwan")
        elif i.split(" ")[1] == "KS":
            region_ordered.append("Korea")
        elif i.split(" ")[1] == "GR":
            region_ordered.append("Germany")
        elif i.split(" ")[1] == "CH":
            region_ordered.append("China")
        elif i.split(" ")[1] == "FR":
            region_ordered.append("France")
        else:
            region_ordered.append("Others")

    theme=paper_portfolio[["Ticker","Theme"]]
    theme=theme.dropna()
    theme[theme.columns[0]]=[i.split(" Equity")[0] for i in theme[theme.columns[0]]]
    GICS=paper_portfolio[["Ticker","GICS Sub-Industry Name"]]

    theme_ordered=[theme[theme[theme.columns[0]]==i][theme.columns[1]].item() if i in list(theme[theme.columns[0]]) else "" for i in (ticker_ytd)]
    GICS_ordered = [GICS[GICS[GICS.columns[0]] == i][GICS.columns[1]].item() if i in list(GICS[GICS.columns[0]]) else "" for i in (ticker_ytd)]
    semi_attribution_itd["Region"]=region_ordered
    semi_attribution_itd["GICS"]=GICS_ordered
    semi_attribution_itd["Theme"]=theme_ordered

    ws = wb["Attribution ITD"]
    ws["A1"].value= "PORT Attribution Report: ("+datetime(2023,6,16).strftime("%Y/%m/%d")+" - "+lastmonth_lastday.strftime("%Y/%m/%d")+")"
    ws["A1"].font=Font(name="Arial",size=12,bold=True,color="000000")
    ws = worksheet(ws, header, semi_attribution_itd,0)
    del wb[wb.sheetnames[0]]
    wb.save("SVLO Global Semi Attribution Summary _"+lastmonth_lastday.strftime("%Y%m%d")+".xlsx")

