import sys
import dateutil
from win32com.client import Dispatch
import win32com.client as win32
from openpyxl.formatting.rule import DataBarRule
import pandas as pd
from openpyxl import load_workbook,Workbook
from datetime import datetime
from openpyxl.styles import *
from datetime import datetime, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from dateutil.relativedelta import relativedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import shutil
import sys
def findUserName():
    path = os.path.expanduser('~')
    pMax = len(path)
    pMin = path.find('Users')+6
    userName = path[pMin:pMax]
    return userName
def web_scrap(url, x_path,renamed_file,file):
    download_dir = 'C:/Users/' + findUserName() + '/Downloads/'
    file_name = file+"_holdings"
    options = webdriver.ChromeOptions()
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option('excludeSwitches', ['enable-automation', 'enable-logging'])
    s = Service('C:/Users/' + findUserName() + '/Documents/chromedriver.exe')
    # driver = webdriver.Chrome(service=s,options=options)
    driver = webdriver.Chrome()
    driver.get(url)
    driver.maximize_window()
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="onetrust-accept-btn-handler"]'))).click()
    # WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="fundHeaderDocLinks"]/li[4]/a'))).click()
    btn = driver.find_element(By.XPATH, x_path)
    driver.execute_script("arguments[0].click();", btn)
    print("Downloading index...")
    downloaded = 0
    timer = 0
    while not downloaded:
        #if(filename[1:] in os.listdir(download_dir)):
        if file_name+".csv" in os.listdir(download_dir):
            downloaded = 1
            break
        time.sleep(1)
        timer += 1
        print(str(timer) + " second(s) have passed")
    if file == "SEMI":
        file_date= pd.read_csv(download_dir+"SEMI_holdings.csv", on_bad_lines='skip', header=None).iloc[0, 1]
        file_date=datetime.strptime(file_date,"%d/%b/%Y")
        shutil.move(download_dir + file_name+".csv", renamed_dir +"\\"+file_name+"_"+file_date.strftime("%Y%m%d")+".csv")
        wb_tickertostrnexcel(file_date,renamed_dir +"\\"+file_name+"_"+file_date.strftime("%Y%m%d"),"SEMI")
        driver.close()
        return renamed_dir +"\\"+file_name+"_"+file_date.strftime("%Y%m%d")+".xlsx", file_date
    else:
        file_date = pd.read_csv(download_dir + "SOXX_holdings.csv", on_bad_lines='skip').iloc[0].item()
        file_date = datetime.strptime(file_date, "%b %d, %Y")
        shutil.move(download_dir + file_name+".csv", renamed_dir +"\\"+file_name+"_"+file_date.strftime("%Y%m%d")+".csv")
        wb_tickertostrnexcel(file_date,renamed_dir +"\\"+file_name+"_"+file_date.strftime("%Y%m%d"),"SOXX")
        os.remove(renamed_dir +"\\"+file_name+"_"+file_date.strftime("%Y%m%d")+".csv")
        driver.close()
def wb_tickertostrnexcel(file_date,file,index):
    if index=="SEMI":
        benchmark= pd.read_csv(file+".csv",on_bad_lines='skip',header =2 ).dropna(axis=0)
        benchmark["Ticker"]=[str(i) for i in benchmark["Ticker"]]
        new_file = Workbook()
        ws = new_file[new_file.sheetnames[0]]
        ws.title = "SEMI_holdings"+file_date.strftime("%m%d")
        ws["A1"].value = "Fund Holdings as of"
        ws["B1"].value = file_date.strftime("%d-%b-%y")
    else:
        benchmark= pd.read_csv(file+".csv",on_bad_lines='skip',header =9 ).dropna(axis=0)
        benchmark["Ticker"]=[str(i) for i in benchmark["Ticker"]]
        new_file = Workbook()
        ws = new_file[new_file.sheetnames[0]]
        ws.title = "SOXX_holdings"+file_date.strftime("%m%d")
        ws["A1"].value = "Fund Holdings as of"
        ws["B1"].value = file_date.strftime("%d-%b-%y")
    rows = dataframe_to_rows(benchmark, index=False, header=True)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    new_file.save(file+".xlsx")

def formatting():
    global font_bold,font_notbold,border,purpleFill,yellowFill,orangeFill,no_fill,no_border,GrayFill,GreenFill,rule_databar
    no_fill = PatternFill(fill_type=None)
    side = Side(border_style=None)
    no_border = borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    font_bold=Font(name="Arial",size=11,bold=True,color="000000")
    font_notbold=Font(name="Arial",size=11,bold=False,color="000000")
    thin=Side(border_style="thin", color="000000")
    double=Side(border_style="double", color="000000")
    border=Border(top=thin, left=thin, right=thin, bottom=thin)

    yellow="FFFF99"  ##YELLOW
    yellowFill=PatternFill(start_color=yellow,
                             end_color=yellow,
                             fill_type='solid')

    purple='9999FF'
    purpleFill=PatternFill(start_color=purple,
                             end_color=purple,
                             fill_type='solid')

    orange="FFC125"  ##orange
    OrangeFill=PatternFill(start_color=orange,
                             end_color=orange,
                             fill_type='solid')
    gray="E7E6E6"  ##gray
    GrayFill=PatternFill(start_color=gray,
                             end_color=gray,
                             fill_type='solid')
    green="B8C2AD"  ##green

    GreenFill=PatternFill(start_color=green,
                             end_color=green,
                             fill_type='solid')
    rule_databar = DataBarRule(start_type="percentile", end_type="percentile", color="63C384", showValue="None", minLength=None,maxLength=None)
    rule_databar = DataBarRule(start_type="percentile",start_value=10, end_type="percentile",
        end_value="90", color="FF638EC6", showValue="None", minLength=None,maxLength=None)


def last_friday():
    today = datetime.today().replace(second=0,hour=0,microsecond=0,minute=0)
    offset = (today.weekday() - 4) % 7
    last_friday_date = today - timedelta(days=offset)
    return last_friday_date
def check_if_updated(date):
    lastfriday= last_friday()
    if date!= lastfriday:
        print("The date shown in the file is "+date.strftime("%d/%b/%Y"))
        print("Please run the code again later today")
        time.sleep(10)
        return 0
    else:
        return 1

def run_macro(macro_wb, macro_name, result_wb):
    xl = win32.gencache.EnsureDispatch("Excel.Application")
    xl.Visible = True
    xl.DisplayAlerts = False
    wb_macro = openWorkbook(xl,macro_wb)
    wb = openWorkbook(xl,result_wb)
    xl.Application.Run(macro_name)
    wb.Close(SaveChanges=True)
    wb_macro.Close()
    xl.Application.Quit
    del xl
def openWorkbook(xlapp, xlfile):
    try:
        xlwb = xlapp.Workbooks(xlfile)
    except Exception as e:
        try:
            xlwb = xlapp.Workbooks.Open(xlfile)
        except Exception as e:
            print(e)
            xlwb = None
    return(xlwb)
if __name__ == "__main__":
    formatting()
    xls="/html/body/div[1]/div[2]/div/div/div/div/div/div[1]/div/div[2]/header[2]/div[1]/div[2]/ul/li[4]/a"
    csv="/html/body/div[1]/div[2]/div/div/div/div/div/div[17]/div/div/div/div[2]/a"
    x_path = '/html/body/div[1]/div[2]/div/div/div/div/div/div[13]/div/div/div/div[2]/a[1]'
    x_url = 'https://www.ishares.com/uk/professional/en/products/319084/ishares-msci-global-semiconductors-ucits-etf?switchLocale=y&siteEntryPassthrough=true'
    renamed_dir = os.getcwd()
    file_name, file_date= web_scrap(x_url, csv, renamed_dir,"SEMI")
    soxx_url="https://www.ishares.com/us/products/239705/ishares-phlx-semiconductor-etf"
    soxx_csv = "/html/body/div[1]/div[2]/div/div/div/div/div/div[14]/div/div/div/div[2]/a"
    correct_date=check_if_updated(file_date)
    print(correct_date)
    if correct_date:
        web_scrap(soxx_url, soxx_csv, renamed_dir,"SOXX")
        wb=load_workbook("SVLO Semi Paper Portfolio_"+(last_friday()-timedelta(7)).strftime("%Y%m%d")+".xlsx")
        ws=wb["Portfolio Weighting"]
        #the function is hided 
        ws=portfolio_weighting(ws)
        ws=wb["Benchmark Weighting"]
        #the function is hided 
        ws=benchmark_weighting(ws,file_name)
        wb.save("SVLO Semi Paper Portfolio_"+last_friday().strftime("%Y%m%d")+".xlsx")
    marco="'"+os.getcwd()+"\\formatting_macro.xlsm'!Module1.databar_format"
    run_macro(os.getcwd()+"\\formatting_macro.xlsm",marco, os.getcwd()+"\\"+"SVLO Semi Paper Portfolio_"+last_friday().strftime("%Y%m%d")+".xlsx")
    os.remove(os.getcwd()+"\\Semi_holdings_"+last_friday().strftime("%Y%m%d")+".csv")
    #os.rename(os.getcwd()+"\\Semi_holdings_"+last_friday().strftime("%Y%m%d")+".excel", os.getcwd()+"\\iShares-MSCI-Global-Semiconductors-UCITS-ETF-USD-Acc_fund-"+last_friday().strftime("%Y%m%d")+".excel")
